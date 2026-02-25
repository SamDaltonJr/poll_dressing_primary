"""
Batch geocode Early Voting locations from the Excel spreadsheet.
Reads all 3 county dispatch route sheets and generates pollingLocations.ts
with geocoded coordinates and tier-based sizing.

Tier mapping: Tier 1 → L (Large), Tier 2 → M (Medium), Tier 3/4 → S (Small)
"""

import sys
import io
import json
import time
import urllib.request
import urllib.parse
import os

# Fix Windows encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

EXCEL_PATH = r"C:\Users\spdal\Downloads\NorthTexas_EV_Talarico_Sign_Distribution (3).xlsx"
OUTPUT_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "src", "config", "pollingLocations.ts")
)

CENSUS_GEOCODER_URL = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"

TIER_TO_SIZE = {
    "Tier 1": "L",
    "Tier 2": "M",
    "Tier 3": "S",
    "Tier 4": "S",
}

COUNTY_PREFIX = {
    "DC": "DC",
    "TC": "TC",
    "DENT": "DN",
}


def geocode_address(address: str):
    """Geocode using US Census Bureau API. Returns (lat, lon) or None."""
    params = urllib.parse.urlencode(
        {"address": address, "benchmark": "Public_AR_Current", "format": "json"}
    )
    url = f"{CENSUS_GEOCODER_URL}?{params}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "BigSignMapper/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
            matches = data.get("result", {}).get("addressMatches", [])
            if matches:
                coords = matches[0]["coordinates"]
                return (round(coords["y"], 5), round(coords["x"], 5))
    except Exception as e:
        print(f"  ERROR: {e}", file=sys.stderr)
    return None


def escape_ts(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def extract_locations(wb):
    """Extract all EV locations from the 3 dispatch route sheets."""
    dispatch_sheets = [s for s in wb.sheetnames if "Dispatch" in s]
    all_locations = []

    for sheet_name in dispatch_sheets:
        ws = wb[sheet_name]
        # Determine county from sheet name prefix
        county = sheet_name.split()[0].replace("—", "").strip()
        if county not in COUNTY_PREFIX:
            # Try matching
            for key in COUNTY_PREFIX:
                if key in sheet_name.upper():
                    county = key
                    break

        # Find header row
        header_row = None
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            vals = [str(v).strip() if v else "" for v in row]
            if "Location" in vals and "Address" in vals:
                header_row = row_idx
                headers = vals
                break

        if not header_row:
            print(f"  Skipping {sheet_name}: no header found")
            continue

        loc_idx = headers.index("Location")
        addr_idx = headers.index("Address")
        city_idx = headers.index("City")
        zip_idx = headers.index("ZIP")
        tier_idx = headers.index("Tier") if "Tier" in headers else -1
        vc_idx = headers.index("VC#") if "VC#" in headers else -1

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            vals = [str(v).strip() if v else "" for v in row]
            location = vals[loc_idx] if loc_idx < len(vals) else ""
            address = vals[addr_idx] if addr_idx < len(vals) else ""
            city = vals[city_idx] if city_idx < len(vals) else ""
            zipcode = vals[zip_idx] if zip_idx < len(vals) else ""
            tier = vals[tier_idx] if tier_idx >= 0 and tier_idx < len(vals) else ""
            vc = vals[vc_idx] if vc_idx >= 0 and vc_idx < len(vals) else ""

            if location and address:
                count += 1
                all_locations.append({
                    "county": county,
                    "vc": vc,
                    "location": location,
                    "address": address,
                    "city": city,
                    "zip": zipcode,
                    "tier": tier,
                })

        print(f"  {sheet_name}: {count} locations")

    return all_locations


def main():
    print(f"Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    locations = extract_locations(wb)
    print(f"\nTotal: {len(locations)} early voting locations to geocode.\n")

    results = []
    failed = []

    for i, loc in enumerate(locations):
        full_address = f"{loc['address']}, {loc['city']}, TX {loc['zip']}"
        print(f"[{i+1}/{len(locations)}] {loc['location']} -- {full_address} ... ", end="", flush=True)

        coords = geocode_address(full_address)
        if coords:
            lat, lon = coords
            print(f"OK ({lat}, {lon})")

            # Generate ID
            county_prefix = COUNTY_PREFIX.get(loc["county"], "EV")
            vc = loc["vc"]
            if vc:
                ev_id = f"ev-{vc}"
            else:
                ev_id = f"ev-{county_prefix}-{i+1}"

            size = TIER_TO_SIZE.get(loc["tier"], "S")

            results.append({
                "id": ev_id,
                "label": loc["location"],
                "address": full_address,
                "latitude": lat,
                "longitude": lon,
                "size": size,
                "tier": loc["tier"],
            })
        else:
            print("FAILED")
            failed.append(loc)

        time.sleep(0.2)

    print(f"\n--- DONE ---")
    print(f"Geocoded: {len(results)} / {len(locations)}")
    print(f"Failed:   {len(failed)}")

    if failed:
        print("\nFailed locations:")
        for f in failed:
            print(f"  {f['location']} -- {f['address']}, {f['city']}, TX {f['zip']}")

    # Write TypeScript file
    print(f"\nWriting: {OUTPUT_PATH}")
    with open(OUTPUT_PATH, "w", encoding="utf-8") as out:
        out.write("import type { MapMarker } from '../types';\n\n")
        out.write("// Early Voting locations across Dallas, Tarrant, and Denton counties\n")
        out.write(f"// Geocoded {len(results)} of {len(locations)} addresses using US Census Bureau API\n")
        out.write("// Tier 1 (300+ Dem ballots) = Large, Tier 2 (150-299) = Medium, Tier 3/4 = Small\n\n")
        out.write("const pollingLocations: MapMarker[] = [\n")

        for r in results:
            out.write("  {\n")
            out.write(f"    id: '{escape_ts(r['id'])}',\n")
            out.write(f"    type: 'earlyVoting',\n")
            out.write(f"    size: '{r['size']}',\n")
            out.write(f"    label: '{escape_ts(r['label'])}',\n")
            out.write(f"    address: '{escape_ts(r['address'])}',\n")
            out.write(f"    latitude: {r['latitude']},\n")
            out.write(f"    longitude: {r['longitude']},\n")
            out.write("  },\n")

        out.write("];\n\n")
        out.write("export default pollingLocations;\n")

    print("Done!")


if __name__ == "__main__":
    main()
